import os
from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime
from io import BytesIO
import cloudinary
import cloudinary.uploader
from PIL import Image
import base64

# === CONFIGURACIÓN INICIAL ===
app = Flask(__name__)
CORS(app)

# Configuración de Cloudinary (pon tus credenciales aquí)
cloudinary.config(
    cloud_name = "dp3xgt5u7", 
    api_key = "425132641747674",
    api_secret="B8OTumlnMQdnM-c_dX_LY7DqkEY",
    secure=True
)

# Ruta del archivo Excel
excel_file = "firmas.xlsx"

# Crear archivo Excel si no existe
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Firmas"
    ws.append(["Identificación", "Fecha y Hora", "URL de Firma"])
    wb.save(excel_file)

# === RUTA PRINCIPAL PARA GUARDAR FIRMA ===
@app.route("/guardar", methods=["POST"])
def guardar():
    identificacion = request.form.get("identificacion")
    imagen_data = request.form.get("firma")

    if not identificacion or not imagen_data:
        return jsonify({"error": "Faltan datos"}), 400

    try:
        # Decodificar imagen base64
        header, encoded = imagen_data.split(",", 1)
        img_bytes = base64.b64decode(encoded)
        img = Image.open(BytesIO(img_bytes))

        # Subir imagen a Cloudinary
        img_io = BytesIO()
        img.save(img_io, 'PNG')
        img_io.seek(0)

        upload_result = cloudinary.uploader.upload(
            img_io,
            public_id=identificacion,
            folder="firmas_app",  # opcional, para organización
            overwrite=True,
            resource_type="image"
        )

        firma_url = upload_result['secure_url']

        # Guardar en Excel
        wb = load_workbook(excel_file)
        sheet = wb.active
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([identificacion, fecha, firma_url])
        wb.save(excel_file)

        return jsonify({"mensaje": "Firma guardada correctamente", "url": firma_url})

    except Exception as e:
        return jsonify({"error": f"Ocurrió un error: {str(e)}"}), 500

# === RUTA PARA DESCARGAR EL EXCEL ===
@app.route("/descargar-excel")
def descargar_excel():
    return send_file(excel_file, as_attachment=True)
  
@app.route("/")
def home():
    return render_template("index.html")

# === INICIO DEL SERVIDOR ===
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))  # Para Render
    app.run(host="0.0.0.0", port=port)